import tkinter as tk
from tkinter import messagebox, ttk, filedialog, colorchooser
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import win32com.client
import os
import time
from PIL import ImageGrab

class EditorHorario:
    def __init__(self, root):
        self.root = root
        self.root.title("Editor de Horario - BUAP - Ingenieria de Software")
        self.root.geometry("1250x750")
        
        self.materias_registradas = []
        self.horarios_temporales = []
        self.dias_semana = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]
        self.color_seleccionado = "#E3F2FD"
        
        self.setup_ui()

    def setup_ui(self):
        self.panel_izq = tk.Frame(self.root, width=400, padx=15, pady=15)
        self.panel_izq.pack(side="left", fill="y")
        
        tk.Label(self.panel_izq, text="REGISTRO DE MATERIA", font=('Arial', 12, 'bold')).pack(pady=10)
        self.ent_nombre = self.crear_campo("Nombre de Materia:")
        self.ent_salon = self.crear_campo("Salón:")
        self.ent_profesor = self.crear_campo("Profesor:")
        
        tk.Label(self.panel_izq, text="Asignar Horario por Día:", font=('Arial', 9, 'bold')).pack(pady=(10,0))
        frame_dia_hora = tk.Frame(self.panel_izq)
        frame_dia_hora.pack(fill="x", pady=5)
        
        self.combo_dia = ttk.Combobox(frame_dia_hora, values=self.dias_semana, width=10, state="readonly")
        self.combo_dia.set("Lunes")
        self.combo_dia.pack(side="left", padx=2)
        
        self.ent_h_inicio = tk.Entry(frame_dia_hora, width=5)
        self.ent_h_inicio.pack(side="left", padx=2)
        tk.Label(frame_dia_hora, text="a").pack(side="left")
        self.ent_h_fin = tk.Entry(frame_dia_hora, width=5)
        self.ent_h_fin.pack(side="left", padx=2)
        
        tk.Button(self.panel_izq, text="+ Agregar Día/Hora", command=self.agregar_horario_temporal).pack(fill="x", pady=2)
        self.lbl_horarios_lista = tk.Label(self.panel_izq, text="Días: (Ninguno)", fg="blue", wraplength=250)
        self.lbl_horarios_lista.pack(fill="x")

        self.btn_color = tk.Button(self.panel_izq, text="Color de Materia", command=self.elegir_color, bg=self.color_seleccionado)
        self.btn_color.pack(fill="x", pady=10)

        tk.Button(self.panel_izq, text="Registrar Materia", bg="#4CAF50", fg="white", 
                  command=self.validar_y_registrar, font=('Arial', 10, 'bold')).pack(fill="x", pady=5)

        tk.Label(self.panel_izq, text="EXPORTAR HORARIO", font=('Arial', 11, 'bold')).pack(pady=(15, 5))
        btn_frame = tk.Frame(self.panel_izq)
        btn_frame.pack(fill="x")
        
        tk.Button(btn_frame, text="PDF", bg="#2196F3", fg="white", 
                  command=lambda: self.exportar_proceso_completo("pdf"), font=('Arial', 10, 'bold')).pack(side="left", expand=True, padx=2)
        
        tk.Button(btn_frame, text="JPG", bg="#FF9800", fg="white", 
                  command=lambda: self.exportar_proceso_completo("jpg"), font=('Arial', 10, 'bold')).pack(side="left", expand=True, padx=2)

        tk.Label(self.panel_izq, text="GESTIONAR VISIBILIDAD", font=('Arial', 11, 'bold')).pack(pady=(15, 5))
        self.listado_materias = tk.Frame(self.panel_izq)
        self.listado_materias.pack(fill="both", expand=True)

        self.tabla_frame = tk.Frame(self.root, bg="white", padx=10, pady=10)
        self.tabla_frame.pack(side="right", expand=True, fill="both")
        self.canvas = tk.Canvas(self.tabla_frame, bg="white", highlightthickness=0)
        self.canvas.pack(expand=True, fill="both")
        self.dibujar_cuadricula()

    def crear_campo(self, texto):
        tk.Label(self.panel_izq, text=texto).pack(anchor="w")
        entry = tk.Entry(self.panel_izq)
        entry.pack(fill="x", pady=2)
        return entry

    def elegir_color(self):
        color = colorchooser.askcolor(title="Seleccionar color")[1]
        if color:
            self.color_seleccionado = color
            self.btn_color.config(bg=color)

    def agregar_horario_temporal(self):
        try:
            dia, h_i, h_f = self.combo_dia.get(), int(self.ent_h_inicio.get()), int(self.ent_h_fin.get())
            if h_i >= h_f: raise ValueError
            self.horarios_temporales.append((dia, h_i, h_f))
            txt = ", ".join([f"{d}({i}-{f})" for d, i, f in self.horarios_temporales])
            self.lbl_horarios_lista.config(text=f"Días: {txt}")
        except:
            messagebox.showerror("Error", "Horas inválidas.")

    def validar_y_registrar(self):
        nombre = self.ent_nombre.get().strip()
        if not nombre or not self.horarios_temporales:
            return messagebox.showerror("Error", "Faltan datos.")

        nueva_materia = {
            "nombre": nombre, "salon": self.ent_salon.get().strip(), "profesor": self.ent_profesor.get().strip(),
            "horarios": list(self.horarios_temporales), "color": self.color_seleccionado,
            "visible": tk.BooleanVar(value=False)
        }
        self.materias_registradas.append(nueva_materia)
        self.limpiar_campos()
        self.actualizar_lista_toggles()
        self.actualizar_tabla()

    def limpiar_campos(self):
        self.ent_nombre.delete(0, tk.END); self.ent_salon.delete(0, tk.END)
        self.ent_profesor.delete(0, tk.END); self.ent_h_inicio.delete(0, tk.END)
        self.ent_h_fin.delete(0, tk.END); self.horarios_temporales = []
        self.lbl_horarios_lista.config(text="Días: (Ninguno)")

    def verificar_choque(self, materia_toggled):
        if materia_toggled['visible'].get():
            for m in self.materias_registradas:
                if m != materia_toggled and m['visible'].get():
                    for dia1, ini1, fin1 in materia_toggled['horarios']:
                        for dia2, ini2, fin2 in m['horarios']:
                            # Nueva validación: Solo hay choque si hay un traslape real de horas.
                            # Ej: 7 a 9 y 9 a 10 no chocan. 7 a 9 y 8 a 10 sí chocan.
                            if dia1 == dia2 and (ini1 < fin2 and fin1 > ini2):
                                messagebox.showwarning(
                                    "Choque de Horario detectado", 
                                    f"La materia '{materia_toggled['nombre']}' se empalma con '{m['nombre']}' el día {dia1}."
                                )
                                materia_toggled['visible'].set(False)
                                self.actualizar_tabla()
                                return
        self.actualizar_tabla()

    def actualizar_lista_toggles(self):
        for w in self.listado_materias.winfo_children(): w.destroy()
        for m in self.materias_registradas:
            f = tk.Frame(self.listado_materias)
            f.pack(fill="x")
            tk.Checkbutton(f, text=f"{m['nombre']} - {m['profesor']}", variable=m['visible'], 
                           command=lambda mat=m: self.verificar_choque(mat)).pack(side="left")

    def dibujar_cuadricula(self):
        self.canvas.delete("all")
        cw, rh = 140, 50
        header_dias = ["Horario"] + self.dias_semana
        for i, dia in enumerate(header_dias):
            x = i * cw
            self.canvas.create_rectangle(x, 0, x + cw, rh, fill="#F8F9FA", outline="#000000")
            self.canvas.create_text(x + cw/2, rh/2, text=dia, font=('Arial', 9, 'bold'))
        
        # El rango ahora es de 7 a 20 (generando bloques de 7:00 a 20:00)
        for j in range(7, 20):
            y = (j - 6) * rh
            self.canvas.create_rectangle(0, y, cw, y + rh, fill="#F8F9FA", outline="#000000")
            # Se agrega el formato de rango en las etiquetas
            self.canvas.create_text(cw/2, y + rh/2, text=f"{j}:00 - {j+1}:00")
            for i in range(1, len(header_dias)):
                self.canvas.create_rectangle(i * cw, y, (i+1) * cw, y + rh, outline="#DEE2E6")

    def actualizar_tabla(self):
        self.canvas.delete("materia")
        cw, rh = 140, 50
        for m in self.materias_registradas:
            if m['visible'].get():
                for dia, h_i, h_f in m['horarios']:
                    if dia in self.dias_semana:
                        col_idx = self.dias_semana.index(dia) + 1
                        y_s, y_e = (h_i - 6) * rh, (h_f - 6) * rh
                        self.canvas.create_rectangle(col_idx * cw, y_s, (col_idx + 1) * cw, y_e, fill=m["color"], outline="black", tags="materia")
                        self.canvas.create_text(col_idx * cw + cw/2, y_s + (y_e - y_s)/2, text=f"{m['nombre']}\n{m['salon']}", justify="center", font=('Arial', 8, 'bold'), tags="materia")

    def exportar_proceso_completo(self, formato):
        materias_visibles = [m for m in self.materias_registradas if m['visible'].get()]
        if not materias_visibles:
            return messagebox.showwarning("Aviso", "No hay materias visibles en el horario.")

        dias_indices = []
        horas = []
        for m in materias_visibles:
            for dia, h_i, h_f in m['horarios']:
                if dia in self.dias_semana:
                    dias_indices.append(self.dias_semana.index(dia))
                    horas.extend([h_i, h_f])

        d_min, d_max = min(dias_indices), max(dias_indices)
        h_min, h_max = min(horas), max(horas)

        wb = openpyxl.Workbook()
        ws = wb.active
        black_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        dias_rango = self.dias_semana[d_min:d_max + 1]

        # --- NUEVO: Dimensiones explícitas de columnas y filas ---
        ws.column_dimensions['A'].width = 15
        for col_idx in range(2, len(dias_rango) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        ws.row_dimensions[1].height = 24
        for row_idx in range(2, (h_max - h_min) + 2):
            ws.row_dimensions[row_idx].height = 40

        # --- Encabezados ---
        c = ws.cell(row=1, column=1, value="HORA")
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = black_border

        for i, dia in enumerate(dias_rango, start=2):
            cell = ws.cell(row=1, column=i, value=dia)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = black_border

        # --- Filas de horas ---
        for i, hora in enumerate(range(h_min, h_max), start=2):
            c = ws.cell(row=i, column=1, value=f"{hora}:00 - {hora+1}:00")
            c.border = black_border
            c.alignment = Alignment(horizontal="center", vertical="center")
            for c_idx in range(2, len(dias_rango) + 2):
                ws.cell(row=i, column=c_idx).border = black_border

        # --- Materias ---
        for m in materias_visibles:
            hex_color = m['color'].lstrip('#')
            m_fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
            for dia, h_i, h_f in m['horarios']:
                if dia in self.dias_semana:
                    col = (self.dias_semana.index(dia) - d_min) + 2
                    row_s = (h_i - h_min) + 2
                    row_e = (h_f - h_min) + 1
                    if row_e >= row_s:
                        ws.merge_cells(start_row=row_s, start_column=col, end_row=row_e, end_column=col)
                        cell = ws.cell(row=row_s, column=col)
                        cell.value = f"{m['nombre']}\n{m['salon']}"
                        cell.fill = m_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.border = black_border

        # --- NUEVO: Configuración de página para que el PDF no recorte ---
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins.left = 0.4
        ws.page_margins.right = 0.4
        ws.page_margins.top = 0.4
        ws.page_margins.bottom = 0.4

        path = os.path.abspath(filedialog.asksaveasfilename(defaultextension=f".{formato}"))
        if not path:
            return

        temp_xlsx = os.path.abspath("temp_export.xlsx")
        wb.save(temp_xlsx)

        try:
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            doc = excel.Workbooks.Open(temp_xlsx)
            ws_sheet = doc.ActiveSheet

            if formato == "pdf":
                # Exporta solo la hoja activa con la configuración de página ya aplicada
                ws_sheet.ExportAsFixedFormat(
                    Type=0,                    # xlTypePDF
                    Filename=path,
                    Quality=0,                 # xlQualityStandard
                    IncludeDocProperties=False,
                    IgnorePrintAreas=False,
                    OpenAfterPublish=False
                )
            else:
                # Excel debe estar visible para que CopyPicture escriba al portapapeles
                excel.Visible = True
                excel.WindowState = -4140  # xlMinimized

                used_range = ws_sheet.UsedRange
                used_range.CopyPicture(Appearance=1, Format=2)  # xlScreen, xlBitmap
                time.sleep(0.8)

                # Leer la imagen directo del portapapeles, sin necesidad del chart
                img = ImageGrab.grabclipboard()
                if img is None:
                    raise Exception("No se pudo capturar la imagen del portapapeles.")
                img.save(path, "JPEG", quality=95)

            doc.Close(False)
            excel.Quit()
            if os.path.exists(temp_xlsx):
                os.remove(temp_xlsx)
            messagebox.showinfo("Éxito", f"Horario guardado correctamente como {formato.upper()}.")
        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error en la exportación: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = EditorHorario(root)
    root.mainloop()